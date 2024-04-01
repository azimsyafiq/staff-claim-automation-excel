import os
import time
import openpyxl
import shutil
import pymsgbox
from datetime import datetime
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

one_drive_path = os.environ.get('OneDriveCommercial')
claim_folder = one_drive_path + r"\Staff Claim"
cash_claim_folder = one_drive_path + r"\Staff Claim\AutoCount Files"
cashbook_file = "CashBook.xlsx"
duplicated_target_file = ""

def copy_data_and_append(source_file, target_file):

    time.sleep(1)
    try:
        # Load the source Excel file
        source_workbook = openpyxl.load_workbook(source_file)
        source_sheet = source_workbook.active

        # Load the target Excel file
        target_workbook = openpyxl.load_workbook(target_file)
        target_sheet = target_workbook.active

        # Find the last non-empty row in column B of source sheet
        last_row_b = source_sheet.max_row
        while source_sheet.cell(row=last_row_b, column=2).value is None and last_row_b > 1:
            last_row_b -= 1

        # Find the last non-empty row in column M of source sheet
        last_row_m = source_sheet.max_row
        while source_sheet.cell(row=last_row_m, column=13).value is None and last_row_m > 1:
            last_row_m -= 1

        # Find the last non-empty row in column V of source sheet
        last_row_v = source_sheet.max_row
        while source_sheet.cell(row=last_row_v, column=22).value is None and last_row_v > 1:
            last_row_v -= 1

        last_row_w = source_sheet.max_row
        while source_sheet.cell(row=last_row_w, column=23).value is None and last_row_w > 1:
            last_row_w -= 1

        last_row_x = source_sheet.max_row
        while source_sheet.cell(row=last_row_x, column=24).value is None and last_row_x > 1:
            last_row_x -= 1

        # Find the last non-empty row in column L of source sheet
        last_row_l = source_sheet.max_row
        while source_sheet.cell(row=last_row_l, column=12).value is None and last_row_l > 1:
            last_row_l -= 1

        # Find the last non-empty row in column O of source sheet
        last_row_o = source_sheet.max_row
        while source_sheet.cell(row=last_row_o, column=15).value is None and last_row_o > 1:
            last_row_o -= 1

        # Find the last non-empty row in column K of source sheet
        last_row_k = source_sheet.max_row
        while source_sheet.cell(row=last_row_k, column=11).value is None and last_row_k > 1:
            last_row_k -= 1

        # Find the last non-empty row in column F of source sheet
        last_row_f = source_sheet.max_row
        while source_sheet.cell(row=last_row_f, column=6).value is None and last_row_f > 1:
            last_row_f -= 1

        # Determine the target start rows in cashclaim.xlsx
        target_start_row_f = target_sheet.max_row + 1
        target_start_row_o = target_sheet.max_row + 1
        target_start_row_q = target_sheet.max_row + 1
        target_start_row_t = target_sheet.max_row + 1
        target_start_row_ad = target_sheet.max_row + 1
        target_start_row_al = target_sheet.max_row + 1
        target_start_row_c = target_sheet.max_row + 1
        target_start_row_g = target_sheet.max_row + 1

        # Copy data from source sheet to target sheet
        for row in range(2, last_row_b + 1):
            source_value_b = source_sheet.cell(row=row, column=2).value
            target_sheet.cell(row=target_start_row_f, column=2, value="<<New>>")
            target_sheet.cell(row=target_start_row_f, column=4, value="PV")
            target_sheet.cell(row=target_start_row_f, column=6, value=source_value_b)
            # target_sheet.cell(row=target_start_row_f, column=7, value="Cash Payment")
            target_sheet.cell(row=target_start_row_f, column=8, value="MYR")
            target_sheet.cell(row=target_start_row_f, column=9, value=1)
            target_sheet.cell(row=target_start_row_f, column=10, value=1)
            target_sheet.cell(row=target_start_row_f, column=16, value=1)
            target_sheet.cell(row=target_start_row_f, column=36, value="CIMB BANK (MYR)")
            target_sheet.cell(row=target_start_row_f, column=40, value=1)
            target_start_row_f += 1

        for row in range(2, last_row_m + 1):
            source_value_m = source_sheet.cell(row=row, column=13).value
            target_sheet.cell(row=target_start_row_o, column=15, value=source_value_m)
            target_start_row_o += 1

        # Copy data from columns V, W, and X of source sheet to column Q of target sheet
        for row in range(2, last_row_v + 1):
            source_value_v = source_sheet.cell(row=row, column=22).value
            source_value_w = source_sheet.cell(row=row, column=23).value
            source_value_x = source_sheet.cell(row=row, column=24).value
            source_value_y = source_sheet.cell(row=row, column=25).value

            # Check if any one of the four columns has a non-empty value
            # if source_value_v != "" or source_value_w != "" or source_value_x != "" or source_value_y != "":
            if any([source_value_v, source_value_w, source_value_x, source_value_y]):
                # Use the first non-empty value among columns V, W, X, and Y
                # if source_value_v != "":
                #     value_to_copy = source_value_v
                # elif source_value_w != "":
                #     value_to_copy = source_value_w
                # elif source_value_x != "":
                #     value_to_copy = source_value_x
                # else:
                #     value_to_copy = source_value_y
                value_to_copy = next((val for val in [source_value_v, source_value_w, source_value_x, source_value_y] if val), "")
                
                target_sheet.cell(row=target_start_row_q, column=17, value=value_to_copy)
                target_start_row_q += 1
            else:
                # If all three columns are empty, paste an empty string to the target cell
                target_sheet.cell(row=target_start_row_q, column=17, value="")
                target_start_row_q += 1

            ###############

        for row in range(2, last_row_l + 1):
            source_value_l = source_sheet.cell(row=row, column=12).value
            target_sheet.cell(row=target_start_row_t, column=20, value=source_value_l)
            target_start_row_t += 1

        for row in range(2, last_row_o + 1):
            source_value_o = source_sheet.cell(row=row, column=15).value
            target_sheet.cell(row=target_start_row_ad, column=30, value=source_value_o)
            target_start_row_ad += 1

        for row in range(2, last_row_o + 1):
            source_value_o = source_sheet.cell(row=row, column=15).value
            target_sheet.cell(row=target_start_row_al, column=38, value=source_value_o)
            target_start_row_al += 1
            
        for row in range(2, last_row_l + 1):
            source_value_f = source_sheet.cell(row=row, column=6).value
            target_sheet.cell(row=target_start_row_g, column=7, value=source_value_f)
            target_start_row_g += 1

        for row in range(2, last_row_k + 1):
            source_value_k = source_sheet.cell(row=row, column=11).value
            
            # Change date format from yyyy-mm-dd to dd/mm/yyyy
            if isinstance(source_value_k, str) and len(source_value_k) == 10 and source_value_k[4] == '-' and source_value_k[7] == '-':
                try:
                    # Parse the string as a datetime object
                    date_obj = datetime.strptime(source_value_k, "%Y-%m-%d")
                    # Format the date as dd/mm/yyyy
                    formatted_date = date_obj.strftime("%d/%m/%Y")
                    target_sheet.cell(row=target_start_row_c, column=3, value=formatted_date)

                except ValueError:
                    target_sheet.cell(row=target_start_row_c, column=3, value=source_value_k)

            else:
                target_sheet.cell(row=target_start_row_c, column=3, value=source_value_k)

            target_start_row_c += 1

        target_workbook.save(target_file)

    except Exception as e:
        print(f"Error copying and pasting data: {str(e)}")

def duplicate_workbook(source_file, target_file):
    global duplicated_target_file

    time.sleep(1)
    try:
        source_workbook = openpyxl.load_workbook(source_file)
        source_sheet = source_workbook.active

        # Get the value of cell B2
        cell_to_duplicate = source_sheet['J2'].value

        # Duplicate the target workbook file
        target_folder = os.path.dirname(target_file)
        # target_filename = os.path.basename(target_file)
        target_filename = f"Cashbook-{cell_to_duplicate}.xlsx"
        duplicated_target_file = os.path.join(target_folder, target_filename) #os.path.join(target_folder, f"{target_filename}-{cell_to_duplicate}")
        shutil.copy2(target_file, duplicated_target_file)

    except Exception as e:
        print(f"Error in duplication: {str(e)}")

# Function to process the Excel file
def process_excel(source_file):
    global duplicated_target_file
    cashbook = os.path.join(cash_claim_folder, cashbook_file)

    duplicate_workbook(source_file, cashbook)

    target_file = duplicated_target_file

    copy_data_and_append(source_file, target_file)

    time.sleep(1)
    try:
        # Load the Excel file
        workbook = openpyxl.load_workbook(source_file)
        sheet = workbook.active

        # Get the value of cell B2
        cell_j2_value = sheet['J2'].value

        # Rename the file with cell B2 value
        new_filename = f"claim-list-{cell_j2_value}.xlsx"
        new_filepath = os.path.join(claim_folder, new_filename)
        os.rename(source_file, new_filepath)

    except Exception as e:
        print(f"Error processing Excel file: {str(e)}")
    
    message = "Claim File Copied for AutoCount:\n{}".format(new_filename)
    pymsgbox.alert(message, "Claim Automation", "OK")

# Define an event handler to watch for file creation events
class ClaimFolderHandler(FileSystemEventHandler):
    def on_created(self, event):
        if not event.is_directory:
            file_path = event.src_path
            if os.path.basename(file_path) == "claim-list.xlsx":
                process_excel(file_path)

if __name__ == "__main__":
    # Create the claim and cash claim folders if they don't exist
    os.makedirs(claim_folder, exist_ok=True)
    os.makedirs(cash_claim_folder, exist_ok=True)

    # Create an observer to watch the claim folder
    observer = Observer()
    observer.schedule(ClaimFolderHandler(), path=claim_folder, recursive=False)
    observer.start()

    try:
        print("Monitoring 'Staff Claim' folder for new files. Press Ctrl+C to exit.")
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("Exiting in")
        for i in range(3, 0, -1):
            print(i)
            time.sleep(1)
        print("Exiting")

        observer.stop()
        observer.join()
